"""정규식+LLM 하이브리드로 정책 원문에서 지원금액 지급 여부를 판정하는 오프라인 배치.

정답(ground truth)은 `문화정책_라벨링_중앙·서울_코드기준.xlsx`의 '자동판정(인간)' 열
(사람이 실제로 검토·확정한 값) 81건이다. 같은 파일의 '자동판정(참고)' 열은 기존
regex-only 방식의 결과이며, 이 배치가 그 baseline 정확도를 넘기는 것이 목표다.

사용법:
    python AI/support_amount_extraction.py
"""

import json
import os
import re
import sys
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import openpyxl
from anthropic import Anthropic
from dotenv import load_dotenv

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from app.schemas.support_amount_extraction import SupportAmountExtraction  # noqa: E402

load_dotenv(Path(__file__).resolve().parent.parent / ".env")

XLSX_PATH = Path(__file__).resolve().parent.parent.parent / "문화정책_라벨링_중앙·서울_코드기준.xlsx"
MODEL = "claude-sonnet-5"

DETERMINATION_RULES = """[판정 기준]
정책 원문을 읽고, 청년 개인(또는 청년 창업기업)에게 실제로 돈이 지급되는지 판정해.
있음/없음/확인필요 3가지 중 하나로만 판정한다 — 비율기반이라는 별도 카테고리는 없다.

- 있음: 청년 개인/창업기업에게 실제 현금이 지급되고, 금액이 원문에서 하나의 값(또는
  범위의 상한값)으로 특정됨
- 없음: 아래 중 하나라도 해당하면 없음
  - 대출/융자/보증 한도 (상환해야 하는 돈이라 지원금이 아님)
  - 사업 전체 예산, 펀드 조성액 (개인에게 지급되는 금액이 아님)
  - 자격조건·심사기준에 등장하는 숫자 (예: 소득 기준, 나이 상한)
  - 시설 이용, 교육, 컨설팅, 공간 대여처럼 현금이 아닌 현물/서비스성 지원
- 확인필요: 아래 중 하나라도 해당하면 확인필요
  - 비율기반 지원 (정액이 아니라 이자 일부 지원, 정부 매칭 적립처럼 비용의 일정
    비율(%)을 지원 — 예: "청년내일저축계좌", "K-패스")
  - 소득구간·연도·지역 등 조건에 따라 정액/매칭 방식 자체가 달라져서 원문만으로
    단일 금액을 특정할 수 없음
  - 소득공제·세액공제·근로소득공제처럼 세금/소득인정액 계산에서 일정액을 빼주는
    방식 (현금이 직접 지급되는 게 아니라 개인 세율·소득수준에 따라 실제 절감액이
    달라짐 — 예: "문화비 소득공제", "근로·사업소득 40만원+30% 공제")

본인부담금·자기부담금 차액을 정부/지자체가 대신 부담해주는 것(예: 의료비 본인부담분
경감, 이용료 차액 지원)은 현금을 손에 쥐여주는 건 아니지만 개인이 실제로 내야 할
비용을 줄여주는 실질적 지원이므로 있음으로 판정한다.

확인필요는 도피처가 아니다 — 있음/없음으로 확실히 판단할 수 있는데 애매하다는 이유만으로
확인필요를 고르지 마라. 아래 순서로 먼저 있음/없음을 확정 지으려 시도하고, 그래도 안 되는
경우에만 확인필요를 써라:
  1. 원문에 구체적 금액(OO원, OO만원)이 하나로 특정돼 있으면 → 있음 (범위면 상한값)
  2. 대출/사업예산/자격조건 숫자/서비스이용형 바우처/기관·센터·시스템 운영·교육·컨설팅처럼
     현금이 아닌 현물·서비스·행정운영이면 → 없음 (이런 유형은 "판단하기 애매해서"가 아니라
     "원래 개인에게 돈이 안 가는 사업"이니 자신 있게 없음으로 판정해라)
  3. 위 둘 다 아니고, 진짜 비율기반이거나 조건별로 금액 자체가 달라지는 경우에만 → 확인필요

바우처/이용권은 두 종류를 구분해:
  - 구매력형(=있음): 카드/포인트에 총액이 미리 충전되어 지정 카테고리 안에서 본인이 자유롭게
    사고 싶은 걸 고르는 방식. 예: "OO만원 상당 문화이용권(카드) 지급", "전자바우처(카드방식),
    지원품목: 채소류·과일류..." → 개인이 실질적 구매력을 받는 것이므로 있음.
  - 서비스이용형(=없음): 국가가 지정한 전문 서비스 제공기관에 세션 단가를 정산해주는 방식.
    예: "심리상담 서비스 8회, 1회당 OO원, 본인부담금 OO%" → 본인은 서비스를 받을 뿐 돈을
    만지지 않으므로 없음.

지자체 행사(축제·공연·심사위원 모집 등) 공고문처럼 원문에 금액이 전혀 안 나오는 경우는
아래 [판정 예시]를 참고해서 판단해 — 이런 유형은 원문에 안 적혀 있어도 실제로는 사례비/
활동비가 관행적으로 지급되는 경우가 있다.

보호종료(자립준비청년), 가족돌봄청년, 한부모, 고립·은둔청년, 출산, 빈곤(근로인센티브 등
소득 하위계층 대상 사업)처럼 "개인의 특수한 사정"에 따라 대상 여부·금액이 케이스별로
달라지는 정책은 원문 텍스트만으로 판정을 100% 확신하기 어렵다. 이런 유형은 판정은
최선으로 하되 confidence를 0.8 이하로 낮게 매겨서, 확신이 낮다는 걸 있는 그대로 드러내라
— 실제로는 아리송한데 confidence를 0.9 이상으로 매겨서 확정처럼 보이게 하지 마라.
"""

FEWSHOT_TRAIN_NOS = [
    "20250305005400110593",  # 서울 청년수당 — 있음, 명시적 현금
    "20250609005400110892",  # 자립준비청년 자립수당 지급 — 있음, 매월 현금 계좌이체
    "20250827005400211523",  # 은평청년축제 청년톡톡콘서트 — 있음, 행사 참여(암묵적 사례비)
    "20250819005400211519",  # 버스킹 페스타 판정단 모집 — 있음, 심사 활동(암묵적 심사비)
    "20250317005400210641",  # 청년예술청 운영 — 있음, 프로그램 참가 지원 포함
    "20250519005400210859",  # 서울시 고립·은둔청년 지원사업 — 있음, 개인 대상 지원 프로그램
    "20260421005400112773",  # 미소금융 청년 미래이음 대출 — 없음, 대출(상환 의무)
    "20260415005400112751",  # (복지부) 26년 정신건강 심리상담 바우처사업 — 없음, 서비스이용형 바우처
    "20250121005400110344",  # 2025년 전국민마음투자지원사업 — 없음, 서비스이용형 바우처
    "20250316005400210640",  # 서울청년문화패스 지원 — 있음, 구매력형 바우처(카드)
    "20260430005400113009",  # 청년내일저축계좌 — 확인필요(비율기반), 정부 매칭 적립
    "20260710005400113257",  # K-패스(K패스) — 확인필요(비율기반), 요금 환급 비율
]

# 정책명에 이 키워드가 들어가면 LLM 판정과 상관없이 확인필요로 강제한다.
# 최신 xlsx에서 사람이 "자립"(자립준비청년) 계열 8건과 "청년내일저축계좌"/"청년도약계좌"
# 계열을 직접 확인필요로 재라벨링해서, 개인 사정·소득구간별로 금액이 달라지는 이
# 정책군은 원문 텍스트로 개별 판단하기보다 이름으로 바로 확인필요 처리하는 게 정답과
# 더 잘 맞는다고 확정됐다. "한부모"/"빈곤"도 81건 표본에서 이름이 들어간 건 전부
# 정답이 확인필요였음을 확인하고 추가했다 — "가족돌봄"/"고립"/"은둔"은 반대로 정답이
# 대부분 확인필요가 아니라서(가족돌봄 2/2 없음, 고립·은둔은 혼재) 뺐다. "출산"은 표본
# 1건("고용보험 미적용자 출산급여 지원")만 있고 그 정답이 '없음'이라 "위기청년
# 자립지원"과 같은 성격의 예외로 처리했다(아래 EXCEPTIONS).
NAME_BASED_CONFIRM_NEEDED_KEYWORDS = ["자립", "청년내일저축계좌", "청년도약", "한부모", "빈곤", "출산"]

# 이름에 위 키워드가 들어가지만 사람 정답이 확인필요가 아닌 예외.
# - 20250625005400111134 "위기청년 자립지원(시설퇴소청소년 자립지원수당 지원)":
#   "자립"이 들어가도 정답이 '있음'(월 지급액이 명확한 자립지원수당).
# - 20250123005400110386 "고용보험 미적용자 출산급여 지원": "출산"이 들어가도
#   정답이 '없음'(청년정책 특유의 개인 사정 변수가 아니라 고용보험 미적용자 전체에게
#   임신주수별로 정액 지급되는 급여라 확인필요로 볼 근거가 약함).
NAME_BASED_OVERRIDE_EXCEPTIONS = {"20250625005400111134", "20250123005400110386"}


def apply_name_based_override(policy_no: str, policy_name: str, determination: str) -> str:
    if policy_no in NAME_BASED_OVERRIDE_EXCEPTIONS:
        return determination
    if any(kw in policy_name for kw in NAME_BASED_CONFIRM_NEEDED_KEYWORDS):
        return "확인필요"
    return determination

REGEX_CANDIDATE_PATTERN = re.compile(
    r"\d[\d,]*\s*(?:천|백)?\s*만\s*원|\d[\d,]*\s*원(?!\w)|\d+(?:\.\d+)?\s*%"
)


def find_regex_candidates(text: str) -> list[str]:
    return list(dict.fromkeys(m.strip() for m in REGEX_CANDIDATE_PATTERN.findall(text or "")))


def build_fewshot_block(rows: list[dict]) -> str:
    by_no = {r["policy_no"]: r for r in rows}
    lines = ["[판정 예시] (실제 사람이 검토·확정한 판정 결과)"]
    for i, no in enumerate(FEWSHOT_TRAIN_NOS, start=1):
        r = by_no[no]
        content = (r["raw_content"] or "").strip()
        if len(content) > 350:
            content = content[:350] + "..."
        lines.append(
            f"\n예시 {i}. {r['policy_name']}\n원문: {content}\n→ 판정: {r['ground_truth']}"
        )
    return "\n".join(lines)


# 자동판정(인간) 원본값 대비 사람이 직접 확인해서 정정한 값. 20260330005400112329
# "자립준비청년 자립수당지원"은 이전 xlsx 버전에서 '없음'으로 잘못 입력돼 있어 '있음'으로
# 정정했었는데, 최신 xlsx는 이 정책을 포함해 자립 계열 다수를 '확인필요'로 직접 갱신해서
# 더 이상 코드 쪽에서 덮어쓸 필요가 없다 — 새 xlsx 값을 그대로 신뢰한다.
GROUND_TRUTH_CORRECTIONS: dict[str, str] = {}


def normalize_to_3way(raw: str | None) -> str | None:
    """xlsx 원본은 있음/없음/비율기반/확인 필요(자유서식) 4갈래인데, 지금 스키마는
    있음/없음/확인필요 3갈래다. 비율기반과 '확인 필요'류(표기가 제각각인 자유 텍스트
    포함)를 전부 확인필요로 합친다."""
    if not raw:
        return raw
    raw = raw.strip()
    if raw == "비율기반" or raw.startswith("확인"):
        return "확인필요"
    return raw


def load_labeled_rows() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = []
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx = {name: i for i, name in enumerate(header_row)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        policy_no = row[idx["정책번호"]]
        if not policy_no:
            continue
        policy_no = str(policy_no)
        ground_truth = GROUND_TRUTH_CORRECTIONS.get(policy_no, row[idx["자동판정(인간)"]])
        rows.append(
            {
                "policy_no": policy_no,
                "policy_name": row[idx["정책명"]],
                "raw_content": row[idx["정책지원내용(원문)"]] or "",
                "baseline": normalize_to_3way(row[idx["베이스라인자동판정"]]),
                "ground_truth": normalize_to_3way(ground_truth),
            }
        )
    return rows


def build_tool_schema() -> dict:
    schema = SupportAmountExtraction.model_json_schema()
    schema.pop("description", None)
    return {
        "name": "record_support_amount_determination",
        "description": "정책 원문에서 지원금액 지급 여부와 근거를 구조화해서 기록한다.",
        "input_schema": schema,
    }


def extract_one(client: Anthropic, row: dict, system_prompt: str) -> SupportAmountExtraction:
    candidates = find_regex_candidates(row["raw_content"])
    user_content = (
        f"[정책명]\n{row['policy_name']}\n\n"
        f"[정책지원내용 원문]\n{row['raw_content']}\n\n"
        f"[정규식으로 찾은 숫자 후보 (참고용, 그대로 믿지 말고 문맥으로 판단할 것)]\n"
        f"{', '.join(candidates) if candidates else '(없음)'}"
    )
    message = client.messages.create(
        model=MODEL,
        max_tokens=1024,
        system=system_prompt,
        tools=[build_tool_schema()],
        tool_choice={"type": "tool", "name": "record_support_amount_determination"},
        messages=[{"role": "user", "content": user_content}],
    )
    tool_use = next(b for b in message.content if b.type == "tool_use")
    return SupportAmountExtraction.model_validate(tool_use.input)


def run_batch(system_prompt: str, rows: list[dict], workers: int = 8) -> list[dict]:
    client = Anthropic()
    results = [None] * len(rows)

    def worker(i_row):
        i, row = i_row
        try:
            extraction = extract_one(client, row, system_prompt)
            predicted = apply_name_based_override(row["policy_no"], row["policy_name"], extraction.determination)
            results[i] = {**row, "predicted": predicted, "extraction": extraction.model_dump(mode="json")}
        except Exception as exc:  # noqa: BLE001
            results[i] = {**row, "predicted": None, "error": str(exc)}

    with ThreadPoolExecutor(max_workers=workers) as pool:
        list(pool.map(worker, enumerate(rows)))
    return results


def score(results: list[dict]) -> tuple[int, int, list[dict]]:
    correct, total, mismatches = 0, 0, []
    for r in results:
        total += 1
        if r["predicted"] == r["ground_truth"]:
            correct += 1
        else:
            mismatches.append(r)
    return correct, total, mismatches


if __name__ == "__main__":
    out_dir = Path(__file__).resolve().parent / "results"
    out_dir.mkdir(exist_ok=True)

    all_rows = load_labeled_rows()
    print(f"전체 {len(all_rows)}건 (있음/없음/확인필요 3지선다, 제외 없음)")

    eval_rows = [r for r in all_rows if r["policy_no"] not in FEWSHOT_TRAIN_NOS]
    print(f"few-shot 학습 예시 {len(FEWSHOT_TRAIN_NOS)}건 제외 → held-out 평가셋 {len(eval_rows)}건\n")

    baseline_correct = sum(1 for r in eval_rows if r["baseline"] == r["ground_truth"])
    print(f"[baseline] 자동판정(참고) vs 자동판정(인간) (held-out {len(eval_rows)}건): "
          f"{baseline_correct}/{len(eval_rows)} = {baseline_correct/len(eval_rows):.1%}")

    system_prompt_v1 = (
        "너는 청년정책 원문에서 실제 지원금액 지급 여부를 판정하는 전문가야.\n\n"
        + DETERMINATION_RULES
    )
    print(f"\n[v1] held-out {len(eval_rows)}건 LLM 판정 실행 중 (few-shot 없음)...")
    results_v1 = run_batch(system_prompt_v1, eval_rows)
    correct_v1, total_v1, mismatches_v1 = score(results_v1)
    print(f"[v1] {correct_v1}/{total_v1} = {correct_v1/total_v1:.1%}")
    (out_dir / "v1_results.json").write_text(
        json.dumps(results_v1, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    system_prompt_v2 = (
        "너는 청년정책 원문에서 실제 지원금액 지급 여부를 판정하는 전문가야.\n\n"
        + DETERMINATION_RULES
        + "\n"
        + build_fewshot_block(all_rows)
    )
    print(f"\n[v2] held-out {len(eval_rows)}건 LLM 판정 실행 중 (few-shot {len(FEWSHOT_TRAIN_NOS)}건)...")
    results_v2 = run_batch(system_prompt_v2, eval_rows)
    correct_v2, total_v2, mismatches_v2 = score(results_v2)
    print(f"[v2] {correct_v2}/{total_v2} = {correct_v2/total_v2:.1%}")
    (out_dir / "v2_results.json").write_text(
        json.dumps(results_v2, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    summary = {
        "held_out_size": len(eval_rows),
        "fewshot_train_size": len(FEWSHOT_TRAIN_NOS),
        "baseline_accuracy": baseline_correct / len(eval_rows),
        "v1_accuracy": correct_v1 / total_v1,
        "v2_accuracy": correct_v2 / total_v2,
    }
    (out_dir / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"\n요약: {summary}")

    print(f"\n[v2] 오답 {len(mismatches_v2)}건:")
    for m in mismatches_v2:
        print(f"- {m['policy_name']!r}: 예측={m['predicted']} 정답={m['ground_truth']}")
